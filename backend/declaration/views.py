# data_collection/views.py
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Declaration, DataDictionary
from .serializers import DeclarationSerializer
import os, io
from django.db.models import Q
from django.conf import settings
from django.core.files.storage import default_storage, FileSystemStorage
from django.core.files.base import ContentFile
from django.core.exceptions import MultipleObjectsReturned
import pandas as pd
import numpy as np
from scipy import stats as scipy_stats
import logging
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError
import magic  # You'll need to install python-magic: pip install python-magic
from django.utils import timezone

logger = logging.getLogger(__name__)

class DeclarationViewSet(viewsets.ModelViewSet):
    queryset = Declaration.objects.all()
    serializer_class = DeclarationSerializer
    
    def create(self, request, *args, **kwargs):
        files = request.FILES
        first_line_is_not_header = request.POST.get('first_line_is_not_header') == 'true'
        first_sheet_has_not_dataset = request.POST.get('first_sheet_has_not_dataset') == 'true'
        column_separator = request.POST.get('column_separator', 'semicolon')
        merge_column_wise = request.POST.get('merge_column_wise') == 'true'

        if not files:
            return Response({"error": "No files provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            dataframes = []
            for file_key in files:
                file = files[file_key]
                file_content = file.read()
                
                if file.name.lower().endswith('.csv'):
                    df = pd.read_csv(io.BytesIO(file_content), header=None if first_line_is_not_header else 0, sep=self.get_separator(column_separator))
                elif file.name.lower().endswith(('.xls', '.xlsx')):
                    if first_sheet_has_not_dataset:
                        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True)
                        sheet_to_read = wb.sheetnames[1] if len(wb.sheetnames) > 1 else wb.sheetnames[0]
                        df = pd.read_excel(io.BytesIO(file_content), sheet_name=sheet_to_read, header=None if first_line_is_not_header else 0, engine='openpyxl')
                    else:
                        df = pd.read_excel(io.BytesIO(file_content), header=None if first_line_is_not_header else 0, engine='openpyxl')
                else:
                    return Response({"error": f"Unsupported file format: {file.name}"}, status=status.HTTP_400_BAD_REQUEST)

                if first_line_is_not_header:
                    df.columns = [f'Col_{i+1}' for i in range(len(df.columns))]

                dataframes.append(df)

            # Merge dataframes based on user choice
            if len(dataframes) > 1:
                if merge_column_wise:
                    # Merge column-wise (horizontally)
                    merged_df = pd.concat(dataframes, axis=1)
                    # Rename duplicate columns
                    merged_df.columns = self.rename_duplicate_columns(merged_df.columns)
                else:
                    # Merge row-wise (vertically)
                    merged_df = pd.concat(dataframes, axis=0, ignore_index=True)
            else:
                merged_df = dataframes[0]

            # Save merged DataFrame
            merged_file_name = f'merged_data_{timezone.now().strftime("%Y%m%d%H%M%S")}.csv'
            merged_file_path = os.path.join('data_files', merged_file_name)
            full_merged_path = os.path.join(settings.MEDIA_ROOT, merged_file_path)
            
            # Ensure the directory exists
            os.makedirs(os.path.dirname(full_merged_path), exist_ok=True)
            
            merged_df.to_csv(full_merged_path, index=False)

            # Create Declaration instance
            declaration = Declaration.objects.create(
                file=merged_file_path,
                name=merged_file_name,
                original_name=merged_file_name
            )

            serializer = self.get_serializer(declaration)
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

        except Exception as e:
            import traceback
            print(traceback.format_exc())  # This will print the full traceback
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def rename_duplicate_columns(self, columns):
        new_columns = []
        seen = set()
        for item in columns:
            counter = 1
            new_item = item
            while new_item in seen:
                new_item = f"{item}_{counter}"
                counter += 1
            new_columns.append(new_item)
            seen.add(new_item)
        return new_columns

    def get_separator(self, separator_name):
        separators = {
            'semicolon': ';',
            'comma': ',',
            'tab': '\t',
            'space': ' '
        }
        return separators.get(separator_name, ';')
        
    def perform_create(self, serializer):
        file = self.request.FILES['file']
        serializer.save(original_name=file.name)

    @action(detail=False, methods=['get'], url_path='by-name/(?P<file_name>.+)')
    def get_by_name(self, request, file_name=None):
        try:
            data_file = Declaration.objects.get(file__endswith=file_name)
            serializer = self.get_serializer(data_file)
            return Response(serializer.data)
        except Declaration.DoesNotExist:
            return Response({"error": f"File '{file_name}' not found."}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['get'])
    def preview(self, request, pk=None):
        data_file = self.get_object()
        file_path = data_file.file.path

        # Ensure the file exists
        if not os.path.exists(file_path):
            return Response({"error": "File not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check file type
        mime = magic.Magic(mime=True)
        file_type = mime.from_file(file_path)
        
        print(f"File path: {file_path}")
        print(f"Detected MIME type: {file_type}")

        try:
            if file_type == 'text/csv' or file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path)
            elif file_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel'] or file_path.lower().endswith(('.xls', '.xlsx')):
                # Try different engines
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                except Exception as openpyxl_error:
                    print(f"Error with openpyxl: {str(openpyxl_error)}")
                    try:
                        df = pd.read_excel(file_path, engine='xlrd')
                    except Exception as xlrd_error:
                        print(f"Error with xlrd: {str(xlrd_error)}")
                        raise ValueError("Unable to read Excel file with available engines.")
            else:
                return Response({"error": f"Unsupported file format: {file_type}"}, status=status.HTTP_400_BAD_REQUEST)

            df = df.replace({np.nan: None})

            preview = {
                "file_name": data_file.name,
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "columns": df.columns.tolist(),
                "top_rows": df.head(5).to_dict(orient='records'),
                "bottom_rows": df.tail(5).to_dict(orient='records'),
                "data_types": df.dtypes.apply(lambda x: x.name).to_dict(),
                "non_null_counts": df.count().to_dict(),
                "first_line_is_not_header": any(col.startswith('Col_') for col in df.columns)
            }

            return Response(preview)
        except Exception as e:
            # Log the full error for debugging
            import traceback
            print(f"Error in preview: {str(e)}")
            print(traceback.format_exc())
            return Response({"error": f"Error reading file: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def calculate_descriptive_stats(self, column_data, data_type, level_of_measurement):
        desc_stats = {}
        if level_of_measurement in ['continuous', 'cardinal']:
            numeric_data = pd.to_numeric(column_data, errors='coerce')
            desc_stats = {
                'Mean': round(numeric_data.mean(), 2),
                'Min': round(numeric_data.min(), 2),
                '1st_Quantile': round(numeric_data.quantile(0.01), 2),
                '5th_Quantile': round(numeric_data.quantile(0.05), 2),
                '25th_Q1': round(numeric_data.quantile(0.25), 2),
                '50th_Median': round(numeric_data.median(), 2),
                '75th_Q3': round(numeric_data.quantile(0.75), 2),
                '95th_Quantile': round(numeric_data.quantile(0.95), 2),
                '99th_Quantile': round(numeric_data.quantile(0.99), 2),
                'Max': round(numeric_data.max(), 2),
                'Std': round(numeric_data.std(), 2),
                'Skewness': round(scipy_stats.skew(numeric_data.dropna()), 2),
                'Kurtosis': round(scipy_stats.kurtosis(numeric_data.dropna()), 2)
            }
        elif level_of_measurement in ['nominal', 'ordinal']:
            value_counts = column_data.value_counts(dropna=False)
            total_count = len(column_data)
            desc_stats = {
                '#_of_Categories': len(value_counts),
                'Mode_Value': value_counts.index[0] if len(value_counts) > 0 else None,
                'Mode_Ratio': round((value_counts.iloc[0] / total_count) * 100, 2) if len(value_counts) > 0 else 0,
                'Missing_Ratio': round((column_data.isnull().sum() / total_count) * 100, 2),
                '#_of_Outlier_Categories': sum((value_counts / total_count) < 0.005)
            }
        return desc_stats

    # to json serialization of NaNs
    def replace_nan_with_none(self, obj):
        if isinstance(obj, dict):
            return {key: self.replace_nan_with_none(value) for key, value in obj.items()}
        elif isinstance(obj, list):
            return [self.replace_nan_with_none(item) for item in obj]
        elif isinstance(obj, float) and np.isnan(obj):
            return None
        return obj
        
    @action(detail=True, methods=['post'])
    def data_dictionary(self, request, pk=None):
        data_file = self.get_object()
        file_path = data_file.file.path
        dictionary_file = request.FILES.get('dictionary')

        # Check file type
        mime = magic.Magic(mime=True)
        file_type = mime.from_file(file_path)
        
        print(f"Data file path: {file_path}")
        print(f"Detected MIME type: {file_type}")

        try:
            if file_type == 'text/csv' or file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path)
            elif file_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel'] or file_path.lower().endswith(('.xls', '.xlsx')):
                # Try different engines
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                except Exception as openpyxl_error:
                    print(f"Error with openpyxl: {str(openpyxl_error)}")
                    try:
                        df = pd.read_excel(file_path, engine='xlrd')
                    except Exception as xlrd_error:
                        print(f"Error with xlrd: {str(xlrd_error)}")
                        raise ValueError("Unable to read Excel file with available engines.")
            else:
                return Response({"error": f"Unsupported file format: {file_type}"}, status=status.HTTP_400_BAD_REQUEST)

            def determine_level_of_measurement(column_data, data_type, unique_count):
                if unique_count == len(column_data):
                    return 'id'
                elif (data_type == 'float64')&(unique_count>1000):
                    return 'continuous'
                elif (data_type == 'float64')&(unique_count<=1000):
                    return 'cardinal'
                elif data_type == 'integer':
                    if unique_count > 1000 or unique_count / len(column_data) > 0.1:
                        return 'continuous'
                    else:
                        if unique_count>5:
                            return 'cardinal'
                        else:
                            return 'nominal'
                elif data_type == 'object':
                    try:
                        pd.to_datetime(column_data, errors='raise', format='%d/%m/%Y %I:%M:%S %p')
                        return 'datetime'
                    except:
                        return 'nominal'
                else:
                    return 'unknown'

            data_dict = []
            for column in df.columns:
                column_data = df[column]
                numeric_data = pd.to_numeric(column_data, errors='coerce')
                
                if numeric_data.notna().all():
                    if all(numeric_data.astype(float) == numeric_data.astype(int)):
                        data_type = 'integer'
                    else:
                        data_type = 'float'
                else:
                    data_type = column_data.dtype.name

                unique_count = column_data.nunique(dropna=False)
                level_of_measurement = determine_level_of_measurement(column_data, data_type, unique_count)

                # Calculate Missing_Ratio and Mode_Ratio
                missing_ratio = (column_data.isnull().sum() / len(column_data)) * 100
                mode_count = column_data.value_counts().iloc[0] if len(column_data) > 0 else 0
                mode_ratio = (mode_count / len(column_data)) * 100

                # Determine Model_Usage_YN
                model_usage_yn = ('No' if level_of_measurement in ['id', 'date', 'datetime', 'timestamp'] or 
                                column in ['Target'] or 'date' in data_type.lower() or 
                                (unique_count/len(column_data))>0.90 
                                else 'Yes')

                # Calculate descriptive statistics
                descriptive_stats = self.calculate_descriptive_stats(column_data, data_type, level_of_measurement)

                data_dict.append({
                    'Feature_Name': column,
                    'Data_Type': data_type,
                    '#_of_Unique_Value': unique_count,
                    'Level_of_Measurement': level_of_measurement,
                    'Missing_Ratio': round(missing_ratio, 2),
                    'Mode_Ratio': round(mode_ratio, 2),
                    'Model_Usage_YN': model_usage_yn,
                    'Descriptive_Stats': descriptive_stats
                })

            # Process dictionary file if provided
            if dictionary_file:
                dict_file_type = mime.from_buffer(dictionary_file.read())
                dictionary_file.seek(0)  # Reset file pointer
                print(f"Dictionary file type: {dict_file_type}")

                try:
                    if dict_file_type == 'text/csv' or dictionary_file.name.lower().endswith('.csv'):
                        dict_df = pd.read_csv(dictionary_file)
                    elif dict_file_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'application/vnd.ms-excel'] or dictionary_file.name.lower().endswith(('.xls', '.xlsx')):
                        dict_df = pd.read_excel(dictionary_file, engine='openpyxl')
                    else:
                        return Response({"error": "Unsupported dictionary file format"}, status=400)

                    dict_df = dict_df.set_index(dict_df.columns[0])
                    for item in data_dict:
                        if item['Feature_Name'] in dict_df.index:
                            description = dict_df.loc[item['Feature_Name'], dict_df.columns[0]]
                            item['Feature_Description'] = description
                            if 'Level_of_Measurement' in dict_df.columns:
                                item['Level_of_Measurement'] = dict_df.loc[item['Feature_Name'], 'Level_of_Measurement']
                            
                            # Save to DataDictionary model
                            DataDictionary.objects.update_or_create(
                                data_file=data_file,
                                column_name=item['Feature_Name'],
                                defaults={'description': description}
                            )

                except Exception as e:
                    print(f"Error processing dictionary file: {str(e)}")
                    return Response({"error": f"Error processing dictionary file: {str(e)}"}, status=400)

            data_dict = self.replace_nan_with_none(data_dict)
            return Response(data_dict)

        except Exception as e:
            import traceback
            print(f"Error in data_dictionary: {str(e)}")
            print(traceback.format_exc())
            return Response({"error": f"Error processing file: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)